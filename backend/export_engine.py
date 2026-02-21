"""
Export / CA Reports Generator
- Excel (primary), CSV, and summary report generation
- P&L, Expense Ledger, GST Summary (India), Multi-currency normalized
"""

import io
import csv
import uuid
import logging
from datetime import datetime, timezone
from typing import List
from motor.motor_asyncio import AsyncIOMotorDatabase

logger = logging.getLogger(__name__)


async def generate_excel_report(db: AsyncIOMotorDatabase, org_id: str, report_type: str,
                                 outlet_id: str = None, date_from: str = None, date_to: str = None) -> bytes:
    """Generate Excel report. Returns bytes for download."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    if report_type == "pnl":
        await _build_pnl_sheet(wb, db, org_id, outlet_id, date_from, date_to)
    elif report_type == "expense_ledger":
        await _build_expense_ledger(wb, db, org_id, outlet_id, date_from, date_to)
    elif report_type == "gst_summary":
        await _build_gst_summary(wb, db, org_id, outlet_id, date_from, date_to)
    elif report_type == "multi_currency":
        await _build_multi_currency_report(wb, db, org_id, outlet_id, date_from, date_to)
    else:
        await _build_pnl_sheet(wb, db, org_id, outlet_id, date_from, date_to)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def generate_csv_report(db: AsyncIOMotorDatabase, org_id: str, report_type: str,
                               outlet_id: str = None, date_from: str = None, date_to: str = None) -> str:
    """Generate CSV report. Returns string."""
    output = io.StringIO()
    writer = csv.writer(output)

    query = {"org_id": org_id}
    if outlet_id:
        query["outlet_id"] = outlet_id
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["document_date"] = date_q

    if report_type == "expense_ledger":
        writer.writerow(["Date", "Supplier", "Type", "Invoice#", "Currency", "Amount", "Tax", "Total", "INR Amount", "Status"])
        docs = await db.documents.find(query, {"_id": 0, "file_base64": 0}).sort("document_date", -1).to_list(5000)
        for d in docs:
            ed = d.get("extracted_data", {})
            writer.writerow([
                d.get("document_date", ""), d.get("supplier_name", ""),
                d.get("document_type", ""), ed.get("invoice_number", ""),
                ed.get("currency", "INR"), ed.get("subtotal", 0),
                ed.get("tax_amount", 0), ed.get("total_amount", 0),
                d.get("converted_inr_amount", 0), d.get("status", ""),
            ])
    else:
        writer.writerow(["Date", "Outlet", "Revenue", "Food Cost", "Food Cost %", "Labor", "Other Expenses", "Profit", "Orders"])
        metrics = await db.daily_metrics.find(query, {"_id": 0}).sort("document_date", -1).to_list(5000)
        outlets_map = {}
        outlets_list = await db.outlets.find({"org_id": org_id}, {"_id": 0}).to_list(100)
        for o in outlets_list:
            outlets_map[o["id"]] = o["name"]
        for m in metrics:
            writer.writerow([
                m.get("document_date", ""), outlets_map.get(m.get("outlet_id", ""), ""),
                m.get("revenue", 0), m.get("food_cost", 0), m.get("food_cost_pct", 0),
                m.get("labor_cost", 0), m.get("other_expenses", 0),
                m.get("profit_estimate", 0), m.get("order_count", 0),
            ])

    return output.getvalue()


async def _build_pnl_sheet(wb, db, org_id, outlet_id, date_from, date_to):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = "P&L Statement"

    # Header styling
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    bold_font = Font(bold=True, size=11)

    query = {"org_id": org_id}
    if outlet_id:
        query["outlet_id"] = outlet_id
    if date_from or date_to:
        dq = {}
        if date_from:
            dq["$gte"] = date_from
        if date_to:
            dq["$lte"] = date_to
        query["document_date"] = dq

    metrics = await db.daily_metrics.find(query, {"_id": 0}).to_list(5000)
    outlets_list = await db.outlets.find({"org_id": org_id}, {"_id": 0}).to_list(100)
    outlets_map = {o["id"]: o["name"] for o in outlets_list}

    # Aggregate by outlet
    outlet_data = {}
    for m in metrics:
        oid = m.get("outlet_id", "unknown")
        if oid not in outlet_data:
            outlet_data[oid] = {"revenue": 0, "food_cost": 0, "labor_cost": 0, "other": 0, "profit": 0}
        outlet_data[oid]["revenue"] += m.get("revenue", 0)
        outlet_data[oid]["food_cost"] += m.get("food_cost", 0)
        outlet_data[oid]["labor_cost"] += m.get("labor_cost", 0)
        outlet_data[oid]["other"] += m.get("other_expenses", 0)
        outlet_data[oid]["profit"] += m.get("profit_estimate", 0)

    # Title
    ws.append(["P&L Statement"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([f"Period: {date_from or 'All'} to {date_to or 'All'}"])
    ws.append([])

    # Headers
    headers = ["Outlet", "Revenue", "Food Cost", "Food Cost %", "Labor Cost", "Other Expenses", "Net Profit", "Profit Margin %"]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    grand_total = {"revenue": 0, "food_cost": 0, "labor_cost": 0, "other": 0, "profit": 0}
    for oid, data in outlet_data.items():
        name = outlets_map.get(oid, oid[:8])
        fc_pct = round(data["food_cost"] / data["revenue"] * 100, 1) if data["revenue"] > 0 else 0
        pm_pct = round(data["profit"] / data["revenue"] * 100, 1) if data["revenue"] > 0 else 0
        ws.append([name, round(data["revenue"], 2), round(data["food_cost"], 2), fc_pct,
                    round(data["labor_cost"], 2), round(data["other"], 2), round(data["profit"], 2), pm_pct])
        for k in grand_total:
            grand_total[k] += data[k]

    # Grand total
    ws.append([])
    total_fc = round(grand_total["food_cost"] / grand_total["revenue"] * 100, 1) if grand_total["revenue"] > 0 else 0
    total_pm = round(grand_total["profit"] / grand_total["revenue"] * 100, 1) if grand_total["revenue"] > 0 else 0
    total_row = ["TOTAL", round(grand_total["revenue"], 2), round(grand_total["food_cost"], 2), total_fc,
                  round(grand_total["labor_cost"], 2), round(grand_total["other"], 2), round(grand_total["profit"], 2), total_pm]
    ws.append(total_row)
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).font = bold_font

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)


async def _build_expense_ledger(wb, db, org_id, outlet_id, date_from, date_to):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = "Expense Ledger"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")

    query = {"org_id": org_id, "status": "processed"}
    if outlet_id:
        query["outlet_id"] = outlet_id
    if date_from or date_to:
        dq = {}
        if date_from:
            dq["$gte"] = date_from
        if date_to:
            dq["$lte"] = date_to
        query["document_date"] = dq

    docs = await db.documents.find(query, {"_id": 0, "file_base64": 0}).sort("document_date", -1).to_list(5000)

    ws.append(["Expense Ledger"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([f"Period: {date_from or 'All'} to {date_to or 'All'}"])
    ws.append([])

    headers = ["Date", "Supplier", "Type", "Invoice#", "Currency", "Subtotal", "Tax", "Total", "INR Amount", "Confidence"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_inr = 0
    for d in docs:
        ed = d.get("extracted_data", {})
        inr = d.get("converted_inr_amount", ed.get("total_amount", 0))
        total_inr += inr
        ws.append([
            d.get("document_date", ""), d.get("supplier_name", ""),
            d.get("document_type", ""), ed.get("invoice_number", ""),
            ed.get("currency", "INR"), ed.get("subtotal", 0), ed.get("tax_amount", 0),
            ed.get("total_amount", 0), inr, f"{d.get('extraction_confidence', 0) * 100:.0f}%",
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "", "", "TOTAL", round(total_inr, 2)])
    ws.cell(row=ws.max_row, column=8).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=9).font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)


async def _build_gst_summary(wb, db, org_id, outlet_id, date_from, date_to):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = "GST Summary"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")

    query = {"org_id": org_id, "status": "processed"}
    if outlet_id:
        query["outlet_id"] = outlet_id
    if date_from or date_to:
        dq = {}
        if date_from:
            dq["$gte"] = date_from
        if date_to:
            dq["$lte"] = date_to
        query["document_date"] = dq

    docs = await db.documents.find(query, {"_id": 0, "file_base64": 0}).to_list(5000)
    suppliers = await db.suppliers.find({"org_id": org_id}, {"_id": 0}).to_list(200)
    supplier_gst = {s["id"]: s.get("gst_id", "N/A") for s in suppliers}

    ws.append(["GST Summary Report (India)"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([f"Period: {date_from or 'All'} to {date_to or 'All'}"])
    ws.append([])

    headers = ["Supplier", "GST No.", "Invoices", "Taxable Amount", "CGST", "SGST", "IGST", "Total Tax", "Total Amount"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # Aggregate by supplier
    supplier_agg = {}
    for d in docs:
        sid = d.get("supplier_id") or d.get("supplier_name", "Unknown")
        if sid not in supplier_agg:
            supplier_agg[sid] = {"name": d.get("supplier_name", "Unknown"), "count": 0, "taxable": 0, "tax": 0, "total": 0}
        ed = d.get("extracted_data", {})
        supplier_agg[sid]["count"] += 1
        supplier_agg[sid]["taxable"] += ed.get("subtotal", 0)
        supplier_agg[sid]["tax"] += ed.get("tax_amount", 0)
        supplier_agg[sid]["total"] += ed.get("total_amount", 0)

    for sid, data in supplier_agg.items():
        gst = supplier_gst.get(sid, "N/A")
        half_tax = round(data["tax"] / 2, 2)
        ws.append([data["name"], gst, data["count"], round(data["taxable"], 2),
                    half_tax, half_tax, 0, round(data["tax"], 2), round(data["total"], 2)])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)


async def _build_multi_currency_report(wb, db, org_id, outlet_id, date_from, date_to):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.active
    ws.title = "Multi-Currency Report"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")

    query = {"org_id": org_id, "status": "processed"}
    if outlet_id:
        query["outlet_id"] = outlet_id
    if date_from or date_to:
        dq = {}
        if date_from:
            dq["$gte"] = date_from
        if date_to:
            dq["$lte"] = date_to
        query["document_date"] = dq

    docs = await db.documents.find(query, {"_id": 0, "file_base64": 0}).sort("document_date", -1).to_list(5000)

    ws.append(["Multi-Currency Normalized Report"])
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([f"Base Currency: INR | Period: {date_from or 'All'} to {date_to or 'All'}"])
    ws.append([])

    headers = ["Date", "Supplier", "Original Currency", "Original Amount", "Exchange Rate", "INR Amount", "Rate Source"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for d in docs:
        ed = d.get("extracted_data", {})
        ws.append([
            d.get("document_date", ""), d.get("supplier_name", ""),
            d.get("original_currency", ed.get("currency", "INR")),
            ed.get("total_amount", 0), d.get("exchange_rate", 1.0),
            d.get("converted_inr_amount", 0), d.get("rate_source", "mock"),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
